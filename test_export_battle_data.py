import unittest
import io
import sys
import os
from openpyxl import load_workbook
from web.app import app, games
from game_logic import Game
from web.config import get_config

class TestExportBattleData(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """Set up test environment once for all tests."""
        # Get test configuration
        config = get_config()
        config.PORT = 5001  # Use a different port for testing
        config.DEBUG = True
        config.TESTING = True
        
        # Configure the app
        app.config.from_object(config)
    
    def setUp(self):
        """Set up test environment before each test."""
        self.app_context = app.app_context()
        self.app_context.push()
        self.client = app.test_client()
        
        # Clear any existing games
        games.clear()
        
        # Create a test game with known initial order
        self.lead_names = ["Lead1", "Lead2", "Lead3", "Lead4"]
        self.follow_names = ["Follow1", "Follow2", "Follow3", "Follow4"]
        self.judge_names = ["Judge1", "Judge2"]
        
        try:
            # Start a new game
            response = self.client.post('/api/start_game', json={
                'leads': ','.join(self.lead_names),
                'follows': ','.join(self.follow_names),
                'judges': ','.join(self.judge_names)
            })
            
            if response.status_code != 200:
                print(f"Error starting game: {response.get_data(as_text=True)}")
                self.fail("Failed to start game")
            
            self.session_id = response.json['session_id']
            
            # Store the initial order from the response
            self.initial_leads = response.json['initial_leads']
            self.initial_follows = response.json['initial_follows']
            
            print(f"Initial leads order: {self.initial_leads}")
            print(f"Initial follows order: {self.initial_follows}")
            
            # Add a test round
            self.add_test_round()
            
        except Exception as e:
            print(f"Error in setUp: {str(e)}")
            self.fail(f"Setup failed: {str(e)}")
    
    def tearDown(self):
        """Clean up after each test."""
        try:
            self.app_context.pop()
            games.clear()
        except Exception as e:
            print(f"Error in tearDown: {str(e)}")
    
    def add_test_round(self):
        """Add a test round with votes."""
        try:
            # Get current pairs
            response = self.client.get(f'/api/get_scores?session_id={self.session_id}')
            if response.status_code != 200:
                print(f"Error getting scores: {response.get_data(as_text=True)}")
                self.fail("Failed to get scores")
            
            game = games[self.session_id]
            
            # Add lead votes
            lead_response = self.client.post('/api/judge_leads', json={
                'session_id': self.session_id,
                'votes': [
                    {'judge': 'Judge1', 'vote': 1},
                    {'judge': 'Judge2', 'vote': 1}
                ],
                'song_info': {
                    'title': 'Test Song',
                    'artist': 'Test Artist',
                    'spotify_url': 'https://spotify.com/test'
                }
            })
            
            if lead_response.status_code != 200:
                print(f"Error adding lead votes: {lead_response.get_data(as_text=True)}")
                self.fail("Failed to add lead votes")
            
            # Add follow votes
            follow_response = self.client.post('/api/judge_follows', json={
                'session_id': self.session_id,
                'votes': [
                    {'judge': 'Judge1', 'vote': 2},
                    {'judge': 'Judge2', 'vote': 2}
                ]
            })
            
            if follow_response.status_code != 200:
                print(f"Error adding follow votes: {follow_response.get_data(as_text=True)}")
                self.fail("Failed to add follow votes")
                
        except Exception as e:
            print(f"Error in add_test_round: {str(e)}")
            self.fail(f"Failed to add test round: {str(e)}")
    
    def test_initial_order_export(self):
        """Test that the initial order is correctly exported in the Excel file."""
        try:
            # Export battle data
            response = self.client.get(f'/api/export_battle_data?session_id={self.session_id}')
            
            # Verify response
            self.assertEqual(response.status_code, 200, 
                           f"Export failed with status {response.status_code}")
            self.assertEqual(response.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           "Incorrect mimetype for Excel file")
            
            # Load the Excel file from the response
            excel_data = io.BytesIO(response.data)
            try:
                wb = load_workbook(excel_data)
            except Exception as e:
                print(f"Error loading Excel file: {str(e)}")
                self.fail("Failed to load Excel file")
            
            # Get the Battle Summary sheet
            self.assertIn("Battle Summary", wb.sheetnames, "Battle Summary sheet not found")
            summary_sheet = wb["Battle Summary"]
            
            # Find the initial order section
            current_row = 7  # Start after "Initial Order" header
            found_initial_order = False
            
            # First, find the "Initial Order" section
            while current_row < summary_sheet.max_row:
                cell_value = summary_sheet[f'A{current_row}'].value
                if cell_value == "Initial Order":
                    found_initial_order = True
                    current_row += 2  # Skip the header and move to content
                    break
                current_row += 1
            
            self.assertTrue(found_initial_order, "Initial Order section not found in Excel file")
            
            # Extract leads from Excel
            excel_leads = []
            while current_row < summary_sheet.max_row:
                lead_entry = summary_sheet[f'A{current_row}'].value
                if not lead_entry or lead_entry == "Follows:":
                    break
                if isinstance(lead_entry, str) and '. ' in lead_entry:
                    lead_name = lead_entry.split('. ', 1)[1]
                    excel_leads.append(lead_name)
                current_row += 1
            
            # Extract follows from Excel
            excel_follows = []
            # Skip the "Follows:" header
            current_row += 1
            while current_row < summary_sheet.max_row:
                follow_entry = summary_sheet[f'A{current_row}'].value
                if not follow_entry or follow_entry == "Final Results":
                    break
                if isinstance(follow_entry, str) and '. ' in follow_entry:
                    follow_name = follow_entry.split('. ', 1)[1]
                    excel_follows.append(follow_name)
                current_row += 1
            
            print(f"Excel leads order: {excel_leads}")
            print(f"Excel follows order: {excel_follows}")
            
            # Verify that the exported order matches the initial order
            self.assertEqual(excel_leads, self.initial_leads, 
                           f"Exported leads order does not match initial order.\nExpected: {self.initial_leads}\nGot: {excel_leads}")
            self.assertEqual(excel_follows, self.initial_follows, 
                           f"Exported follows order does not match initial order.\nExpected: {self.initial_follows}\nGot: {excel_follows}")
            
        except Exception as e:
            print(f"Error in test_initial_order_export: {str(e)}")
            self.fail(f"Test failed: {str(e)}")

if __name__ == '__main__':
    unittest.main() 