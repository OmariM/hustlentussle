from flask import Flask, render_template, request, jsonify, send_file, redirect
import sys
import os
import pandas as pd
import io
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from werkzeug.utils import secure_filename
import random
import requests
from flask_cors import CORS
from dotenv import load_dotenv
import uuid
import tempfile
import shutil
import threading
import time
import urllib.parse
import json
import base64
import logging
import atexit
from typing import Optional, Dict, List

# Load environment variables from .env file
load_dotenv()

# Add parent directory to path to import game_logic and persistence
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from game_logic import Game, Contestant
from web.config import get_config
from persistence import (
    RepositoryFactory,
    CleanupScheduler,
    PersistenceError,
    MemoryGameRepository,
)

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Get configuration based on environment
config = get_config()

app = Flask(__name__, 
            static_folder='.',
            static_url_path='',
            template_folder='.')  # Set template folder to current directory
app.config.from_object(config)

# Cache-busting for static assets (helps ensure browsers pull latest JS/CSS after deploy)
if not app.config.get('ASSET_VERSION'):
    # Prefer deploy-provided versioning when available; fallback to process start time.
    app.config['ASSET_VERSION'] = (
        os.environ.get('ASSET_VERSION')
        or os.environ.get('RENDER_GIT_COMMIT')
        or os.environ.get('GIT_COMMIT')
        or str(int(time.time()))
    )
CORS(app)

# Stronger cache control: ensure index.html and assets revalidate after deploy.
# This prevents "hard to shake" stale JS/CSS in some browsers/CDNs.
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

@app.after_request
def add_cache_control_headers(response):
    try:
        path = request.path or ''
        # Never cache the HTML shell; it is the entry point that references versioned assets.
        if path == '/' or response.mimetype == 'text/html':
            response.headers['Cache-Control'] = 'no-store, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        # For static assets, require revalidation (query-param versioning already busts caches,
        # but this makes refreshes more reliable across intermediaries).
        if path.endswith('.js') or path.endswith('.css'):
            response.headers['Cache-Control'] = 'no-cache, max-age=0, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
    except Exception:
        # Never fail a request due to cache headers.
        pass
    return response


# Initialize game repository using factory
def init_repository():
    """Initialize the game repository based on configuration."""
    database_url = getattr(config, 'DATABASE_URL', None)
    fallback_enabled = getattr(config, 'PERSISTENCE_FALLBACK_ENABLED', True)
    expiration_seconds = getattr(config, 'GAME_EXPIRATION_SECONDS', 6 * 60 * 60)
    
    try:
        repo = RepositoryFactory.create_repository(
            database_url=database_url,
            fallback_enabled=fallback_enabled,
            expiration_seconds=expiration_seconds,
        )
        logger.info(f"Repository initialized: {type(repo).__name__}")
        return repo
    except PersistenceError as e:
        logger.error(f"Failed to initialize repository: {e}")
        if fallback_enabled:
            logger.warning("Falling back to in-memory repository")
            return MemoryGameRepository(expiration_seconds=expiration_seconds)
        raise


repo = init_repository()

# Initialize cleanup scheduler
cleanup_scheduler = CleanupScheduler(
    repository=repo,
    interval_seconds=getattr(config, 'CLEANUP_INTERVAL_SECONDS', 60 * 60),
)
cleanup_scheduler.start()

# Register cleanup on shutdown
@atexit.register
def shutdown_cleanup():
    """Cleanup on application shutdown."""
    # Guard against cleanup_scheduler not being defined
    # (can happen if init_repository() raised an exception)
    if 'cleanup_scheduler' in globals() and cleanup_scheduler is not None:
        logger.info("Shutting down cleanup scheduler...")
        cleanup_scheduler.stop()


# Backwards compatibility for tests that import `games`
# Note: This only works with MemoryGameRepository
# For PostgreSQL, tests should use the repo interface directly
if hasattr(repo, '_games'):
    games = repo._games
else:
    games = {}  # Empty dict as placeholder for non-memory repositories

# In-memory user OAuth token store keyed by provided key (session_id or auth_key)
_spotify_user_tokens: Dict[str, dict] = {}
_spotify_user_tokens_lock = threading.RLock()

def _resolve_key(session_id: Optional[str], auth_key: Optional[str]) -> Optional[str]:
    return auth_key or session_id


def _get_redirect_uri() -> str:
    # Prefer explicit env var; else infer from request
    env_uri = os.getenv('SPOTIFY_REDIRECT_URI')
    if env_uri:
        return env_uri
    try:
        base = request.host_url.rstrip('/')
        return f"{base}/api/spotify/callback"
    except Exception:
        # Fallback localhost
        return "http://localhost:5000/api/spotify/callback"


def _now() -> float:
    return time.time()


def _store_user_tokens(key: str, access_token: str, refresh_token: Optional[str], expires_in: int) -> None:
    if not key:
        return
    with _spotify_user_tokens_lock:
        _spotify_user_tokens[key] = {
            'access_token': access_token,
            'refresh_token': refresh_token,
            'expires_at': _now() + max(1, int(expires_in) - 30)  # refresh a bit early
        }


def _get_user_tokens(key: str) -> Optional[dict]:
    if not key:
        return None
    with _spotify_user_tokens_lock:
        return _spotify_user_tokens.get(key)


def _refresh_user_token(key: str) -> Optional[str]:
    tokens = _get_user_tokens(key)
    if not tokens or not tokens.get('refresh_token'):
        return None
    client_id = os.getenv('SPOTIFY_CLIENT_ID')
    client_secret = os.getenv('SPOTIFY_CLIENT_SECRET')
    if not client_id or not client_secret:
        return None
    resp = requests.post(
        'https://accounts.spotify.com/api/token',
        auth=(client_id, client_secret),
        data={
            'grant_type': 'refresh_token',
            'refresh_token': tokens['refresh_token']
        }
    )
    if resp.status_code != 200:
        return None
    data = resp.json()
    access_token = data.get('access_token')
    expires_in = int(data.get('expires_in', 3600))
    # Spotify may or may not return a new refresh token
    refresh_token = data.get('refresh_token', tokens['refresh_token'])
    _store_user_tokens(key, access_token, refresh_token, expires_in)
    return access_token


def _ensure_user_access_token(key: str) -> Optional[str]:
    tokens = _get_user_tokens(key)
    if not tokens:
        return None
    if tokens.get('expires_at', 0) <= _now():
        return _refresh_user_token(key)
    return tokens.get('access_token')


def serialize_state(game: Game) -> dict:
    """Build a canonical renderable game state snapshot for the UI."""
    # Determine crown status using current game flags/threshold
    def has_earned_crown(contestant: Contestant, role: str) -> bool:
        if role == "lead":
            return contestant.points >= game.win_threshold and game.has_winning_lead
        return contestant.points >= game.win_threshold and game.has_winning_follow

    # Build unique contestant maps including current pairs, queues and tracked winners
    lead_dict: Dict[str, dict] = {}
    follow_dict: Dict[str, dict] = {}

    # Add current pair contestants first
    if game.pair_1 and game.pair_2:
        lead_dict[game.pair_1[0].name] = {
            'name': game.pair_1[0].name,
            'points': game.pair_1[0].points,
            'is_winner': has_earned_crown(game.pair_1[0], 'lead')
        }
        lead_dict[game.pair_2[0].name] = {
            'name': game.pair_2[0].name,
            'points': game.pair_2[0].points,
            'is_winner': has_earned_crown(game.pair_2[0], 'lead')
        }
        follow_dict[game.pair_1[1].name] = {
            'name': game.pair_1[1].name,
            'points': game.pair_1[1].points,
            'is_winner': has_earned_crown(game.pair_1[1], 'follow')
        }
        follow_dict[game.pair_2[1].name] = {
            'name': game.pair_2[1].name,
            'points': game.pair_2[1].points,
            'is_winner': has_earned_crown(game.pair_2[1], 'follow')
        }

    # Add contestants from queues (proxies iterate without current competitors)
    for lead in game.leads:
        lead_dict[lead.name] = {
            'name': lead.name,
            'points': lead.points,
            'is_winner': has_earned_crown(lead, 'lead')
        }
    for follow in game.follows:
        follow_dict[follow.name] = {
            'name': follow.name,
            'points': follow.points,
            'is_winner': has_earned_crown(follow, 'follow')
        }

    # Include tracked winners if present
    if getattr(game, 'winning_lead', None):
        lead_dict[game.winning_lead.name] = {
            'name': game.winning_lead.name,
            'points': game.winning_lead.points,
            'is_winner': has_earned_crown(game.winning_lead, 'lead')
        }
    if getattr(game, 'winning_follow', None):
        follow_dict[game.winning_follow.name] = {
            'name': game.winning_follow.name,
            'points': game.winning_follow.points,
            'is_winner': has_earned_crown(game.winning_follow, 'follow')
        }

    # Sort for stable rendering
    lead_list = sorted(list(lead_dict.values()), key=lambda x: (-x['points'], x['name']))
    follow_list = sorted(list(follow_dict.values()), key=lambda x: (-x['points'], x['name']))

    # Judges
    contestant_judges = [j.name for j in game.contestant_judges]
    contestant_enabled = getattr(game, 'contestant_judging_enabled', True)
    simple_flag = bool(getattr(game, 'simple_contestant_judges', False)) and contestant_enabled

    # Build lightweight rounds summary for live UI (include completed + current)
    rounds_data: List[dict] = []
    try:
        for r in getattr(game, 'rounds', []):
            rounds_data.append({
                'round_num': getattr(r, 'round_num', None),
                'pairs': getattr(r, 'pairs', None),
                'lead_winner': getattr(r, 'lead_winner', None),
                'follow_winner': getattr(r, 'follow_winner', None),
            })
        cr = getattr(game, 'current_round', None)
        if cr and cr not in getattr(game, 'rounds', []):
            rounds_data.append({
                'round_num': getattr(cr, 'round_num', None),
                'pairs': getattr(cr, 'pairs', None),
                'lead_winner': getattr(cr, 'lead_winner', None),
                'follow_winner': getattr(cr, 'follow_winner', None),
            })
        rounds_data = [rd for rd in rounds_data if rd.get('round_num') is not None]
        rounds_data.sort(key=lambda x: x['round_num'])
    except Exception:
        rounds_data = []

    state = {
        'session_id': game.session_id,
        'round': {
            'number': game.round_num,
            'pairs': {
                'pair_1': {'lead': game.pair_1[0].name, 'follow': game.pair_1[1].name},
                'pair_2': {'lead': game.pair_2[0].name, 'follow': game.pair_2[1].name},
            },
            'judges': {
                'guest': game.guest_judges,
                'contestant': contestant_judges,
                'simple_contestant_judges': simple_flag,
                'contestant_judging_enabled': contestant_enabled,
                'num_contestant_judges': getattr(game, 'num_contestant_judges', len(contestant_judges)),
                'expected_contestant_judges': getattr(game, 'expected_contestant_judges', len(game.guest_judges) + 1),
            },
        },
        'scoreboard': {
            'leads': lead_list,
            'follows': follow_list,
        },
        'rounds': rounds_data,
        'thresholds': {
            'win': game.win_threshold,
            'auto_win': getattr(game, 'auto_win_threshold', game.win_threshold),
        },
        'flags': {
            'has_winning_lead': game.has_winning_lead,
            'has_winning_follow': game.has_winning_follow,
            'finished': game.is_finished(),
        },
        'winners': {
            'lead': getattr(game, 'winning_lead', None).name if getattr(game, 'winning_lead', None) else None,
            'follow': getattr(game, 'winning_follow', None).name if getattr(game, 'winning_follow', None) else None,
        },
        'initial_order': {
            'leads': [c.name for c in getattr(game, 'initial_leads', [])],
            'follows': [c.name for c in getattr(game, 'initial_follows', [])],
        },
        'queue_order': {
            # Current competitors (positions 1-2) then queue (positions 3+)
            'leads': (
                [game.pair_1[0].name, game.pair_2[0].name] if game.pair_1 and game.pair_2 else []
            ) + [c.name for c in game.leads],
            'follows': (
                [game.pair_1[1].name, game.pair_2[1].name] if game.pair_1 and game.pair_2 else []
            ) + [c.name for c in game.follows],
        },
    }
    return state


@app.route('/')
def index():
    return render_template('index.html', config=app.config)

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint to verify persistence status."""
    repo_type = type(repo).__name__
    game_count = len(repo.list_sessions()) if hasattr(repo, 'list_sessions') else 'unknown'
    
    return jsonify({
        'status': 'healthy',
        'persistence': {
            'repository_type': repo_type,
            'using_postgres': repo_type == 'PostgresGameRepository',
            'active_games': game_count,
        }
    })

@app.route('/api/state', methods=['GET'])
def get_state():
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400
    game = repo.get(session_id)
    if not game:
        return jsonify({'error': 'Game not found'}), 404
    return jsonify(serialize_state(game))

@app.route('/api/start_game', methods=['POST'])
def start_game():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON data provided'}), 400
    
    lead_names = data.get('leads', '').split(',')
    follow_names = data.get('follows', '').split(',')
    judge_names = data.get('judges', '').split(',')
    playlist_url = data.get('playlist_url', None)
    randomize_order = data.get('randomize_order', True)
    contestant_judging_enabled = bool(data.get('contestant_judging_enabled', True))
    simple_contestant_judges = bool(data.get('simple_contestant_judges', False)) and contestant_judging_enabled
    
    # Filter out any empty names
    lead_names = [name.strip() for name in lead_names if name.strip()]
    follow_names = [name.strip() for name in follow_names if name.strip()]
    judge_names = [name.strip() for name in judge_names if name.strip()]
    
    # Parse points_to_win_mode: "default" (7), "auto" (formula), or numeric value
    points_to_win_mode = data.get('points_to_win_mode', 'default')
    points_to_win = data.get('points_to_win', None)  # Legacy support
    
    if points_to_win is not None:
        # Legacy: direct points_to_win value takes precedence
        try:
            parsed = int(points_to_win)
            if parsed >= 1:
                points_to_win = parsed
            else:
                points_to_win = 7
        except (ValueError, TypeError):
            points_to_win = 7
    elif points_to_win_mode == 'auto':
        # Auto-calculate based on number of contestants
        points_to_win = max(len(lead_names), len(follow_names)) - 1
        if points_to_win < 1:
            points_to_win = 1
    elif points_to_win_mode == 'default':
        points_to_win = 7
    else:
        # Treat as custom numeric value
        try:
            points_to_win = int(points_to_win_mode)
            if points_to_win < 1:
                points_to_win = 7
        except (ValueError, TypeError):
            points_to_win = 7
    
    # Parse num_contestant_judges
    num_contestant_judges_raw = data.get('num_contestant_judges', None)
    num_contestant_judges = None
    contestant_judges_warning = None
    expected_contestant_judges = len(judge_names) + 1
    
    if num_contestant_judges_raw is not None:
        try:
            num_contestant_judges = int(num_contestant_judges_raw)
            if num_contestant_judges < 0:
                num_contestant_judges = 0
            # Check if it violates the recommended rule
            if num_contestant_judges != expected_contestant_judges:
                contestant_judges_warning = (
                    f"Warning: You specified {num_contestant_judges} contestant judge(s), "
                    f"but the recommended number is {expected_contestant_judges} "
                    f"(1 more than the {len(judge_names)} guest judge(s)). "
                    "This may cause undetermined behavior."
                )
        except (ValueError, TypeError):
            num_contestant_judges = None
    
    # Determine whether to randomize order
    if isinstance(randomize_order, str):
        randomize_order = randomize_order.strip().lower() in {"1", "true", "yes", "on"}
    randomize_order = bool(randomize_order)
    
    # Randomize the order of leads and follows if requested
    if randomize_order:
        random.shuffle(lead_names)
        random.shuffle(follow_names)
    
    # Create a new game with the configuration
    game = Game(
        lead_names,
        follow_names,
        judge_names,
        contestant_judging_enabled=contestant_judging_enabled,
        simple_contestant_judges=simple_contestant_judges,
        num_contestant_judges=num_contestant_judges,
        points_to_win=points_to_win,
    )

    session_id = repo.create(game)
    
    # Get initial game state
    state = game.get_game_state()
    
    response = {
        'session_id': session_id,
        'round': state['round'],
        'pair_1': state['pair_1'],
        'pair_2': state['pair_2'],
        'contestant_judges': state['contestant_judges'],
        'guest_judges': game.guest_judges,
        'initial_leads': [c.name for c in game.initial_leads],
        'initial_follows': [c.name for c in game.initial_follows],
        'playlist_url': playlist_url or '',
        'simple_contestant_judges': simple_contestant_judges,
        'contestant_judging_enabled': contestant_judging_enabled,
        'randomize_order': randomize_order,
        'points_to_win': game.win_threshold,
        'auto_win_threshold': game.auto_win_threshold,
        'num_contestant_judges': game.num_contestant_judges,
        'expected_contestant_judges': game.expected_contestant_judges,
    }
    
    # Add warning if contestant judges rule is violated
    if contestant_judges_warning:
        response['contestant_judges_warning'] = contestant_judges_warning
    
    return jsonify(response)

@app.route('/api/get_scores', methods=['GET'])
def get_scores():
    session_id = request.args.get('session_id')
    
    game = repo.get(session_id) if session_id else None
    if not game:
        return jsonify({'error': 'Game not found'}), 404

    # Reuse canonical scoreboard
    canon = serialize_state(game)
    return jsonify({
        'leads': canon['scoreboard']['leads'],
        'follows': canon['scoreboard']['follows']
    })

@app.route('/api/judge_leads', methods=['POST'])
def judge_leads():
    data = request.get_json()
    session_id = data.get('session_id')
    votes = data.get('votes', [])
    song_info = data.get('song_info', {})  # Add song info handling
    
    if not session_id or not votes:
        return jsonify({'error': 'Missing session_id or votes'}), 400
    
    game = repo.get(session_id)
    if not game:
        return jsonify({'error': 'Invalid session ID'}), 400
    
    # Store song info in the current round
    if song_info:
        game.current_round.song_info = song_info
    
    # Process votes and determine winner
    result = game.judge_round(game.pair_1[0], game.pair_2[0], "lead", votes)
    
    # Persist game state changes
    repo.save(session_id, game)
    
    return jsonify({
        'winner': result['winner'],
        'guest_votes': result['guest_votes'],
        'contestant_votes': result['contestant_votes']
    })

@app.route('/api/judge_follows', methods=['POST'])
def judge_follows():
    data = request.get_json()
    session_id = data.get('session_id')
    votes = data.get('votes', [])
    song_info = data.get('song_info', {})  # Add song info handling
    
    if not session_id or not votes:
        return jsonify({'error': 'Missing session_id or votes'}), 400
    
    game = repo.get(session_id)
    if not game:
        return jsonify({'error': 'Invalid session ID'}), 400
    
    # Update song info in the current round if not already set
    if song_info and not hasattr(game.current_round, 'song_info'):
        game.current_round.song_info = song_info
    
    # Process votes and determine winner
    result = game.judge_round(game.pair_1[1], game.pair_2[1], "follow", votes)
    
    # Check for win condition
    win_messages = game.check_for_win() or []
    
    # Persist game state changes
    repo.save(session_id, game)
    
    return jsonify({
        'winner': result['winner'],
        'guest_votes': result['guest_votes'],
        'contestant_votes': result['contestant_votes'],
        'win_messages': win_messages,
        'game_finished': game.is_finished()
    })

@app.route('/api/judge_combined', methods=['POST'])
def judge_combined():
    data = request.get_json()
    session_id = data.get('session_id')
    lead_votes = data.get('lead_votes', [])
    follow_votes = data.get('follow_votes', [])
    song_info = data.get('song_info', {})
    
    if not session_id or not lead_votes or not follow_votes:
        return jsonify({'error': 'Missing session_id, lead_votes, or follow_votes'}), 400
    
    game = repo.get(session_id)
    if not game:
        return jsonify({'error': 'Invalid session ID'}), 400
    
    # Store song info in the current round
    if song_info:
        game.current_round.song_info = song_info
    
    # If simple contestant judges is enabled, aggregate the proxy vote to all contestant judges
    def expand_with_mock_contestant_judges(votes_list: List[tuple]) -> List[tuple]:
        try:
            if not getattr(game, 'contestant_judging_enabled', True):
                return votes_list
            if not getattr(game, 'simple_contestant_judges', False):
                return votes_list
            # Find the proxy vote if present
            proxy_name = 'Contestant Judges'
            proxy_votes = [v for v in votes_list if v[0] == proxy_name]
            if not proxy_votes:
                return votes_list
            proxy_decision = proxy_votes[-1][1]
            # Collect the real contestant judge names from current round
            real_cj = list(getattr(game.current_round, 'contestant_judges', []) or [])
            # Remove existing contestant judge entries
            filtered = [(v, d) for (v, d) in votes_list if (v not in real_cj and v != proxy_name)]
            # Expand to mock judges, numbered deterministically
            expanded = filtered[:]
            for idx, _name in enumerate(real_cj, start=1):
                expanded.append((f"Contestant Judge {idx}", proxy_decision))
            return expanded
        except Exception:
            return votes_list

    # Expand votes when needed
    lead_votes_effective = expand_with_mock_contestant_judges(lead_votes)
    follow_votes_effective = expand_with_mock_contestant_judges(follow_votes)

    # Process lead votes and determine winner
    lead_result = game.judge_round(game.pair_1[0], game.pair_2[0], "lead", lead_votes_effective)
    
    # Process follow votes and determine winner
    follow_result = game.judge_round(game.pair_1[1], game.pair_2[1], "follow", follow_votes_effective)
    
    # Check for win condition
    win_messages = game.check_for_win() or []
    
    # Persist game state changes
    repo.save(session_id, game)
    
    return jsonify({
        'lead_winner': lead_result['winner'],
        'lead_guest_votes': lead_result['guest_votes'],
        'lead_contestant_votes': lead_result['contestant_votes'],
        'follow_winner': follow_result['winner'],
        'follow_guest_votes': follow_result['guest_votes'],
        'follow_contestant_votes': follow_result['contestant_votes'],
        'win_messages': win_messages,
        'game_finished': game.is_finished()
    })

@app.route('/api/next_round', methods=['POST'])
def next_round():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON data provided'}), 400
    
    session_id = data.get('session_id')
    
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400
    
    game = repo.get(session_id)
    if not game:
        return jsonify({'error': 'Invalid session ID'}), 400
    game.next_round()
    
    # Persist game state changes
    repo.save(session_id, game)
    
    state = game.get_game_state()
    
    return jsonify({
        'round': state['round'],
        'pair_1': state['pair_1'],
        'pair_2': state['pair_2'],
        'contestant_judges': state['contestant_judges']
    })

@app.route('/api/end_game', methods=['POST'])
def end_game():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON data provided'}), 400
    
    session_id = data.get('session_id')
    
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400
    
    game = repo.get(session_id)
    if not game:
        return jsonify({'error': 'Invalid session ID'}), 400
    
    # Mark the game as finished so display mode polling detects it
    game.state = 1
    repo.save(session_id, game)
    
    leads, follows = game.finalize_results()
    
    # Format the results - include all leads
    lead_results = []
    # Get all leads from initial order to ensure we include everyone
    all_leads = [lead.name for lead in game.initial_leads]
    # Create a dictionary of lead points for easy lookup
    lead_points = {lead.name: lead.points for lead in leads}
    
    # Sort leads by points (descending) and then by name (ascending) for consistent ordering
    sorted_leads = sorted(all_leads, key=lambda x: (-lead_points.get(x, 0), x))
    
    for idx, lead_name in enumerate(sorted_leads):
        points = lead_points.get(lead_name, 0)
        medal = ["🥇", "🥈", "🥉"][idx] if idx < 3 else ""
        is_winner = hasattr(game, 'last_lead_winner') and game.last_lead_winner == lead_name
        lead_results.append({
            'name': lead_name,
            'points': points,
            'medal': medal,
            'is_winner': is_winner
        })
    
    # Format the results - include all follows
    follow_results = []
    # Get all follows from initial order to ensure we include everyone
    all_follows = [follow.name for follow in game.initial_follows]
    # Create a dictionary of follow points for easy lookup
    follow_points = {follow.name: follow.points for follow in follows}
    
    # Sort follows by points (descending) and then by name (ascending) for consistent ordering
    sorted_follows = sorted(all_follows, key=lambda x: (-follow_points.get(x, 0), x))
    
    for idx, follow_name in enumerate(sorted_follows):
        points = follow_points.get(follow_name, 0)
        medal = ["🥇", "🥈", "🥉"][idx] if idx < 3 else ""
        is_winner = hasattr(game, 'last_follow_winner') and game.last_follow_winner == follow_name
        follow_results.append({
            'name': follow_name,
            'points': points,
            'medal': medal,
            'is_winner': is_winner
        })
    
    # Collect round metadata
    rounds_data = []
    
    # Add all completed rounds that belong to this session
    for r in game.rounds:
        if hasattr(r, 'session_id') and r.session_id == session_id:
            round_data = {
                'round_num': r.round_num,
                'session_id': session_id,
                'pairs': r.pairs,
                'lead_votes': r.lead_votes,
                'follow_votes': r.follow_votes,
                'judges': r.judges,
                'contestant_judges': r.contestant_judges,
                'win_messages': r.win_messages,
                'lead_winner': r.lead_winner,
                'follow_winner': r.follow_winner,
                'song_info': r.song_info if hasattr(r, 'song_info') else None
            }
            rounds_data.append(round_data)
    
    # Also include the current round if it exists and is not already in the rounds list
    if game.current_round and game.current_round not in game.rounds:
        if hasattr(game.current_round, 'session_id') and game.current_round.session_id == session_id:
            current_round_data = {
                'round_num': game.current_round.round_num,
                'session_id': session_id,
                'pairs': game.current_round.pairs,
                'lead_votes': game.current_round.lead_votes,
                'follow_votes': game.current_round.follow_votes,
                'judges': game.current_round.judges,
                'contestant_judges': game.current_round.contestant_judges,
                'win_messages': game.current_round.win_messages,
                'lead_winner': game.current_round.lead_winner,
                'follow_winner': game.current_round.follow_winner,
                'song_info': game.current_round.song_info if hasattr(game.current_round, 'song_info') else None
            }
            rounds_data.append(current_round_data)
    
    # Sort rounds by round number
    rounds_data.sort(key=lambda x: x['round_num'])
    
    return jsonify({
        'session_id': session_id,
        'leads': lead_results,
        'follows': follow_results,
        'rounds': rounds_data,
        'initial_leads': [c.name for c in game.initial_leads],
        'initial_follows': [c.name for c in game.initial_follows]
    })

@app.route('/api/export_battle_data', methods=['GET', 'POST'])
def export_battle_data():
    """Export battle data as an Excel file or JSON."""
    session_id = request.args.get('session_id')
    format_type = request.args.get('format', 'excel')  # Default to excel format
    
    if not repo.exists(session_id):
        return jsonify({'error': 'Game not found'}), 404
    
    game = repo.get(session_id)
    
    # Get all rounds data
    all_rounds = []
    all_rounds.extend(game.rounds)
    if game.current_round not in game.rounds:
        all_rounds.append(game.current_round)
    
    # Format the results
    leads, follows = game.finalize_results()
    
    lead_results = []
    for idx, lead in enumerate(leads):
        medal = ["🥇", "🥈", "🥉"][idx] if idx < 3 else ""
        is_winner = hasattr(game, 'last_lead_winner') and game.last_lead_winner == lead.name
        lead_results.append({
            'name': lead.name,
            'points': lead.points,
            'medal': medal,
            'is_winner': is_winner
        })
    
    follow_results = []
    for idx, follow in enumerate(follows):
        medal = ["🥇", "🥈", "🥉"][idx] if idx < 3 else ""
        is_winner = hasattr(game, 'last_follow_winner') and game.last_follow_winner == follow.name
        follow_results.append({
            'name': follow.name,
            'points': follow.points,
            'medal': medal,
            'is_winner': is_winner
        })
    
    # Collect round metadata
    rounds_data = []
    
    # Check if we received updated rounds data in the POST request
    if request.method == 'POST' and request.json and 'rounds' in request.json:
        rounds_data = request.json['rounds']
    else:
        # Add all completed rounds from the game object
        for r in all_rounds:
            round_data = {
                'round_num': r.round_num,
                'session_id': session_id,
                'pairs': r.pairs,
                'lead_votes': r.lead_votes,
                'follow_votes': r.follow_votes,
                'judges': r.judges,
                'contestant_judges': r.contestant_judges,
                'win_messages': r.win_messages,
                'lead_winner': r.lead_winner,
                'follow_winner': r.follow_winner,
                'song_info': r.song_info if hasattr(r, 'song_info') else None
            }
            rounds_data.append(round_data)
    
    # Sort rounds by round number
    rounds_data.sort(key=lambda x: x['round_num'])
    
    if format_type == 'json':
        return jsonify({
            'session_id': session_id,
            'leads': lead_results,
            'follows': follow_results,
            'rounds': rounds_data
        })
    else:
        # Create Excel file in the previous multi-sheet format
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Battle Summary"

        # Add game ID, date, and time
        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        summary_sheet['A1'] = f"Game ID: {session_id}"
        summary_sheet['A3'] = "Date:"
        summary_sheet['B3'] = date_str
        summary_sheet['A4'] = "Time:"
        summary_sheet['B4'] = time_str
        summary_sheet['A5'] = "Total Rounds:"
        summary_sheet['B5'] = len(game.rounds) + (1 if game.current_round not in game.rounds else 0)

        # Initial Order section
        summary_sheet['A7'] = "Initial Order"
        summary_sheet['A9'] = "Leads:"
        initial_leads = [lead.name for lead in game.initial_leads]
        for i, lead in enumerate(initial_leads, 1):
            summary_sheet[f'A{9+i}'] = f"{i}. {lead}"
        follow_start_row = 9 + len(initial_leads) + 2
        summary_sheet[f'A{follow_start_row}'] = "Follows:"
        initial_follows = [follow.name for follow in game.initial_follows]
        for i, follow in enumerate(initial_follows, 1):
            summary_sheet[f'A{follow_start_row+i}'] = f"{i}. {follow}"

        # Final Results section
        results_start_row = follow_start_row + len(initial_follows) + 2
        summary_sheet[f'A{results_start_row}'] = "Final Results"
        summary_sheet[f'A{results_start_row+2}'] = "Lead Winners:"
        for i, lead in enumerate(leads[:3], 1):
            medal = ["🥇", "🥈", "🥉"][i-1]
            is_winner = hasattr(game, 'last_lead_winner') and game.last_lead_winner == lead.name
            crown = " 👑" if is_winner else ""
            summary_sheet[f'A{results_start_row+2+i}'] = f"{medal} {lead.name}{crown}"
            summary_sheet[f'B{results_start_row+2+i}'] = f"{lead.points} points"
        follow_winners_start = results_start_row + 2 + len(leads[:3]) + 2
        summary_sheet['A' + str(follow_winners_start)] = "Follow Winners:"
        for i, follow in enumerate(follows[:3], 1):
            medal = ["🥇", "🥈", "🥉"][i-1]
            is_winner = hasattr(game, 'last_follow_winner') and game.last_follow_winner == follow.name
            crown = " 👑" if is_winner else ""
            summary_sheet[f'A{follow_winners_start+i}'] = f"{medal} {follow.name}{crown}"
            summary_sheet[f'B{follow_winners_start+i}'] = f"{follow.points} points"

        # Lead Leaderboard
        lead_sheet = wb.create_sheet("Lead Leaderboard")
        lead_sheet['A1'] = "Lead"
        lead_sheet['B1'] = "Points"
        for i, lead in enumerate(leads, 2):
            lead_sheet[f'A{i}'] = lead.name
            lead_sheet[f'B{i}'] = lead.points

        # Follow Leaderboard
        follow_sheet = wb.create_sheet("Follow Leaderboard")
        follow_sheet['A1'] = "Follow"
        follow_sheet['B1'] = "Points"
        for i, follow in enumerate(follows, 2):
            follow_sheet[f'A{i}'] = follow.name
            follow_sheet[f'B{i}'] = follow.points

        # Round History
        round_sheet = wb.create_sheet("Round History")
        base_headers = ["Round", "Lead 1", "Lead 2", "Follow 1", "Follow 2", "Song Title", "Artist", "Spotify Link"]
        # Determine the maximum number of judges in any round
        max_judges_count = 0
        for round_data in rounds_data:
            all_judges = set()
            if 'judges' in round_data:
                all_judges.update(round_data['judges'])
            if 'contestant_judges' in round_data:
                all_judges.update(round_data['contestant_judges'])
            max_judges_count = max(max_judges_count, len(all_judges))
        judge_headers = []
        for i in range(max_judges_count):
            judge_headers.append(f"Judge {i+1}")
            judge_headers.append(f"Lead Vote {i+1}")
            judge_headers.append(f"Follow Vote {i+1}")
        headers = base_headers + judge_headers
        for col, header in enumerate(headers, 1):
            round_sheet.cell(row=1, column=col).value = header
        for i, round_data in enumerate(rounds_data, 2):
            round_sheet.cell(row=i, column=1).value = round_data['round_num']
            if 'pairs' in round_data and round_data['pairs']:
                lead1 = round_data['pairs'].get('pair_1', {}).get('lead', '')
                lead2 = round_data['pairs'].get('pair_2', {}).get('lead', '')
                follow1 = round_data['pairs'].get('pair_1', {}).get('follow', '')
                follow2 = round_data['pairs'].get('pair_2', {}).get('follow', '')
                
                # Add contestants and set red color for winners
                lead1_cell = round_sheet.cell(row=i, column=2)
                lead1_cell.value = lead1
                if lead1 == round_data.get('lead_winner'):
                    lead1_cell.font = Font(color="FF0000")
                
                lead2_cell = round_sheet.cell(row=i, column=3)
                lead2_cell.value = lead2
                if lead2 == round_data.get('lead_winner'):
                    lead2_cell.font = Font(color="FF0000")
                
                follow1_cell = round_sheet.cell(row=i, column=4)
                follow1_cell.value = follow1
                if follow1 == round_data.get('follow_winner'):
                    follow1_cell.font = Font(color="FF0000")
                
                follow2_cell = round_sheet.cell(row=i, column=5)
                follow2_cell.value = follow2
                if follow2 == round_data.get('follow_winner'):
                    follow2_cell.font = Font(color="FF0000")
                
                # Song info
                if 'song_info' in round_data and round_data['song_info']:
                    song_info = round_data['song_info']
                    round_sheet.cell(row=i, column=6).value = song_info.get('title', '')
                    round_sheet.cell(row=i, column=7).value = song_info.get('artist', '')
                    round_sheet.cell(row=i, column=8).value = song_info.get('spotify_url', '')
                # Judges and votes
                all_judges = []
                if 'judges' in round_data:
                    all_judges.extend(round_data['judges'])
                if 'contestant_judges' in round_data:
                    all_judges.extend(round_data['contestant_judges'])
                for j in range(max_judges_count):
                    judge_col = 9 + j * 3
                    lead_vote_col = judge_col + 1
                    follow_vote_col = judge_col + 2
                    judge_name = all_judges[j] if j < len(all_judges) else ""
                    if judge_name:
                        lead_vote = round_data['lead_votes'].get(judge_name, '')
                        follow_vote = round_data['follow_votes'].get(judge_name, '')
                        round_sheet.cell(row=i, column=judge_col).value = judge_name
                        
                        # Set lead vote and color if it matches winner
                        lead_vote_cell = round_sheet.cell(row=i, column=lead_vote_col)
                        # Convert lead vote number to contestant name or special case
                        if lead_vote == 1:
                            lead_vote_cell.value = lead1
                            if lead1 == round_data.get('lead_winner'):
                                lead_vote_cell.font = Font(color="FF0000")
                        elif lead_vote == 2:
                            lead_vote_cell.value = lead2
                            if lead2 == round_data.get('lead_winner'):
                                lead_vote_cell.font = Font(color="FF0000")
                        elif lead_vote == 0:
                            lead_vote_cell.value = "Tie"
                        else:
                            lead_vote_cell.value = "No Contest"
                        
                        # Set follow vote and color if it matches winner
                        follow_vote_cell = round_sheet.cell(row=i, column=follow_vote_col)
                        # Convert follow vote number to contestant name or special case
                        if follow_vote == 1:
                            follow_vote_cell.value = follow1
                            if follow1 == round_data.get('follow_winner'):
                                follow_vote_cell.font = Font(color="FF0000")
                        elif follow_vote == 2:
                            follow_vote_cell.value = follow2
                            if follow2 == round_data.get('follow_winner'):
                                follow_vote_cell.font = Font(color="FF0000")
                        elif follow_vote == 0:
                            follow_vote_cell.value = "Tie"
                        else:
                            follow_vote_cell.value = "No Contest"
        # Voting History
        voting_sheet = wb.create_sheet("Voting History")
        voting_headers = ["Round", "Judge", "Lead Vote", "Follow Vote"]
        for col, header in enumerate(voting_headers, 1):
            voting_sheet.cell(row=1, column=col).value = header
        row_index = 2
        for round_data in rounds_data:
            if 'pairs' in round_data and round_data['pairs']:
                all_judges = set()
                if 'judges' in round_data:
                    all_judges.update(round_data['judges'])
                if 'contestant_judges' in round_data:
                    all_judges.update(round_data['contestant_judges'])
                for judge in all_judges:
                    lead_vote = round_data['lead_votes'].get(judge, '')
                    follow_vote = round_data['follow_votes'].get(judge, '')
                    voting_sheet.cell(row=row_index, column=1).value = round_data['round_num']
                    voting_sheet.cell(row=row_index, column=2).value = judge
                    voting_sheet.cell(row=row_index, column=3).value = lead_vote
                    voting_sheet.cell(row=row_index, column=4).value = follow_vote
                    row_index += 1
        # Save the workbook
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'battle_results_{session_id}.xlsx'
        )

@app.route('/api/process_uploaded_file', methods=['POST'])
def process_uploaded_file():
    """Process an uploaded battle history Excel file and return the data for display."""
    if 'battle_file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
        
    file = request.files['battle_file']
    
    if not file or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
        
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'File must be an Excel (.xlsx) file'}), 400
        
    # Save the file temporarily
    temp_path = os.path.join('/tmp', secure_filename(file.filename))
    file.save(temp_path)
    
    try:
        # Load the Excel workbook
        wb = load_workbook(temp_path)
        
        # Extract data from relevant sheets
        data = {}
        
        # Track all contestants and their points from rounds
        all_contestants = {
            'leads': {},
            'follows': {}
        }
        
        # Battle Summary - Process this first to get initial order
        if "Battle Summary" in wb.sheetnames:
            summary_sheet = wb["Battle Summary"]
            
            # Extract date and time
            date = summary_sheet['B3'].value if 'B3' in summary_sheet else None
            time = summary_sheet['B4'].value if 'B4' in summary_sheet else None
            total_rounds = summary_sheet['B5'].value if 'B5' in summary_sheet else 0
            
            print("Processing Battle Summary sheet...")
            print(f"Date: {date}, Time: {time}, Total Rounds: {total_rounds}")
            
            # Extract initial order
            initial_leads = []
            initial_follows = []
            
            # Find the initial order section
            current_row = 7  # Start after "Initial Order" header
            found_initial_order = False
            
            # First, find the "Initial Order" section
            while current_row < summary_sheet.max_row:
                cell_value = summary_sheet[f'A{current_row}'].value
                print(f"Checking row {current_row}: {cell_value}")
                if cell_value == "Initial Order":
                    found_initial_order = True
                    print("Found Initial Order section")
                    current_row += 2  # Skip the header and move to content
                    break
                current_row += 1
            
            if found_initial_order:
                print("Processing initial order section...")
                # Process leads
                while current_row < summary_sheet.max_row:
                    cell_value = summary_sheet[f'A{current_row}'].value
                    print(f"Processing row {current_row}: {cell_value}")
                    if cell_value == "Leads:":
                        print("Found Leads section")
                        current_row += 1
                        while current_row < summary_sheet.max_row:
                            lead_entry = summary_sheet[f'A{current_row}'].value
                            print(f"Processing lead entry: {lead_entry}")
                            if not lead_entry or lead_entry == "Follows:":
                                break
                            # Extract name from "1. Name" format
                            if isinstance(lead_entry, str) and '. ' in lead_entry:
                                lead_name = lead_entry.split('. ', 1)[1]
                                print(f"Added lead: {lead_name}")
                                initial_leads.append(lead_name)
                            current_row += 1
                    elif cell_value == "Follows:":
                        print("Found Follows section")
                        current_row += 1
                        while current_row < summary_sheet.max_row:
                            follow_entry = summary_sheet[f'A{current_row}'].value
                            print(f"Processing follow entry: {follow_entry}")
                            if not follow_entry or follow_entry == "Final Results":
                                break
                            # Extract name from "1. Name" format
                            if isinstance(follow_entry, str) and '. ' in follow_entry:
                                follow_name = follow_entry.split('. ', 1)[1]
                                print(f"Added follow: {follow_name}")
                                initial_follows.append(follow_name)
                            current_row += 1
                    elif cell_value == "Final Results":
                        print("Found Final Results section")
                        break
                    current_row += 1
            
            print(f"Final initial leads: {initial_leads}")
            print(f"Final initial follows: {initial_follows}")
            
            # Store initial order in data
            data['initial_leads'] = initial_leads
            data['initial_follows'] = initial_follows
            
            # Calculate row numbers for results sections
            results_start_row = current_row  # This is where "Final Results" was found
            follow_winners_start = results_start_row + 2 + len(initial_leads) + 2  # Add 2 for headers and spacing
            
            # Extract top leads
            top_leads = []
            for i in range(results_start_row + 2, results_start_row + 5):  # Rows for top 3 leads
                name_cell = f'A{i}'
                points_cell = f'B{i}'
                if name_cell in summary_sheet and summary_sheet[name_cell].value:
                    name_text = summary_sheet[name_cell].value
                    points_text = summary_sheet[points_cell].value if points_cell in summary_sheet else "0 points"
                    
                    # Parse name and medal
                    parts = name_text.split(' ')
                    medal = parts[0] if parts[0] in ["🥇", "🥈", "🥉"] else ""
                    is_winner = "👑" in name_text
                    name = name_text.replace(medal, '').replace("👑", '').strip()
                    
                    # Parse points
                    points = 0
                    try:
                        if isinstance(points_text, str):
                            points = int(points_text.split(' ')[0])
                        elif isinstance(points_text, (int, float)):
                            points = int(points_text)
                    except (ValueError, TypeError):
                        # If parsing fails, try to get points from our calculated data
                        if name in all_contestants['leads']:
                            points = all_contestants['leads'][name]['points']
                    
                    # Mark as winner in our tracking dict
                    if name in all_contestants['leads'] and is_winner:
                        all_contestants['leads'][name]['is_winner'] = True
                    
                    top_leads.append({
                        'name': name,
                        'points': points,
                        'medal': medal,
                        'is_winner': is_winner
                    })
            
            # Extract top follows
            top_follows = []
            for i in range(follow_winners_start + 1, follow_winners_start + 4):  # Rows for top 3 follows
                name_cell = f'A{i}'
                points_cell = f'B{i}'
                if name_cell in summary_sheet and summary_sheet[name_cell].value:
                    name_text = summary_sheet[name_cell].value
                    points_text = summary_sheet[points_cell].value if points_cell in summary_sheet else "0 points"
                    
                    # Parse name and medal
                    parts = name_text.split(' ')
                    medal = parts[0] if parts[0] in ["🥇", "🥈", "🥉"] else ""
                    is_winner = "👑" in name_text
                    name = name_text.replace(medal, '').replace("👑", '').strip()
                    
                    # Parse points
                    points = 0
                    try:
                        if isinstance(points_text, str):
                            points = int(points_text.split(' ')[0])
                        elif isinstance(points_text, (int, float)):
                            points = int(points_text)
                    except (ValueError, TypeError):
                        # If parsing fails, try to get points from our calculated data
                        if name in all_contestants['follows']:
                            points = all_contestants['follows'][name]['points']
                    
                    # Mark as winner in our tracking dict
                    if name in all_contestants['follows'] and is_winner:
                        all_contestants['follows'][name]['is_winner'] = True
                    
                    top_follows.append({
                        'name': name,
                        'points': points,
                        'medal': medal,
                        'is_winner': is_winner
                    })
            
            data['summary'] = {
                'date': date,
                'time': time,
                'total_rounds': total_rounds
            }
        
        # Round History - Process this first to collect all contestant points
        if "Round History" in wb.sheetnames:
            round_sheet = wb["Round History"]
            rounds = []
            
            # Get column headers to understand the sheet format
            headers = {}
            for col in range(1, round_sheet.max_column + 1):
                header = round_sheet.cell(row=1, column=col).value
                if header:
                    headers[col] = header
            
            # Count how many judges (every 3 columns after column 5)
            judge_count = (round_sheet.max_column - 5) // 3
            
            # Extract round data
            for row in range(2, round_sheet.max_row + 1):
                round_num = round_sheet.cell(row=row, column=1).value
                if round_num:
                    lead1 = round_sheet.cell(row=row, column=2).value
                    lead2 = round_sheet.cell(row=row, column=3).value
                    follow1 = round_sheet.cell(row=row, column=4).value
                    follow2 = round_sheet.cell(row=row, column=5).value
                    
                    # Initialize contestants if they don't exist yet
                    for lead in [lead1, lead2]:
                        if lead and lead not in all_contestants['leads']:
                            all_contestants['leads'][lead] = {'name': lead, 'points': 0, 'is_winner': False}
                            
                    for follow in [follow1, follow2]:
                        if follow and follow not in all_contestants['follows']:
                            all_contestants['follows'][follow] = {'name': follow, 'points': 0, 'is_winner': False}
                    
                    round_data = {
                        'round_num': round_num,
                        'pairs': {
                            'pair_1': {
                                'lead': lead1,
                                'follow': follow1
                            },
                            'pair_2': {
                                'lead': lead2,
                                'follow': follow2
                            }
                        },
                        'lead_votes': {},
                        'follow_votes': {}
                    }
                    
                    # Extract song information if available
                    song_title = round_sheet.cell(row=row, column=6).value
                    artist = round_sheet.cell(row=row, column=7).value
                    spotify_url = round_sheet.cell(row=row, column=8).value
                    
                    if song_title or artist or spotify_url:
                        round_data['song_info'] = {
                            'title': song_title or '',
                            'artist': artist or '',
                            'spotify_url': spotify_url or ''
                        }
                    
                    # Extract winner information and award points
                    lead_winner = None
                    follow_winner = None
                    
                    for col in [2, 3]:  # Lead columns
                        cell = round_sheet.cell(row=row, column=col)
                        font_color = cell.font.color
                        if font_color and font_color.rgb == "FF0000":  # Red color
                            lead_winner = cell.value
                            round_data['lead_winner'] = lead_winner
                            # Award a point to the winner
                            if lead_winner and lead_winner in all_contestants['leads']:
                                all_contestants['leads'][lead_winner]['points'] += 1
                    
                    for col in [4, 5]:  # Follow columns
                        cell = round_sheet.cell(row=row, column=col)
                        font_color = cell.font.color
                        if font_color and font_color.rgb == "FF0000":  # Red color
                            follow_winner = cell.value
                            round_data['follow_winner'] = follow_winner
                            # Award a point to the winner
                            if follow_winner and follow_winner in all_contestants['follows']:
                                all_contestants['follows'][follow_winner]['points'] += 1
                    
                    # Extract judge votes
                    for j in range(judge_count):
                        judge_col = 9 + j * 3  # Start after song info columns (6-8)
                        lead_vote_col = judge_col + 1
                        follow_vote_col = judge_col + 2
                        
                        judge_name = round_sheet.cell(row=row, column=judge_col).value
                        if judge_name:
                            lead_vote = round_sheet.cell(row=row, column=lead_vote_col).value
                            follow_vote = round_sheet.cell(row=row, column=follow_vote_col).value
                            
                            if lead_vote:
                                round_data['lead_votes'][judge_name] = lead_vote
                            if follow_vote:
                                round_data['follow_votes'][judge_name] = follow_vote
                    
                    rounds.append(round_data)
            
            data['rounds'] = rounds
        
        # Lead and Follow Leaderboards
        if "Lead Leaderboard" in wb.sheetnames:
            lead_sheet = wb["Lead Leaderboard"]
            leads = []
            for row in range(2, lead_sheet.max_row + 1):
                name = lead_sheet.cell(row=row, column=1).value
                points_cell = lead_sheet.cell(row=row, column=2).value
                if name:
                    # Try to parse points from the cell
                    points = 0
                    try:
                        if isinstance(points_cell, (int, float)):
                            points = int(points_cell)
                        elif isinstance(points_cell, str) and points_cell.isdigit():
                            points = int(points_cell)
                    except (ValueError, TypeError):
                        # If parsing fails, try to get points from our calculated data
                        if name in all_contestants['leads']:
                            points = all_contestants['leads'][name]['points']
                            
                    # Check if this contestant is a winner
                    is_winner = False
                    if name in all_contestants['leads']:
                        is_winner = all_contestants['leads'][name]['is_winner']
                    
                    leads.append({
                        'name': name,
                        'points': points,
                        'is_winner': is_winner
                    })
                    
                    # Update our tracking dictionary
                    if name in all_contestants['leads']:
                        all_contestants['leads'][name]['points'] = max(all_contestants['leads'][name]['points'], points)
            data['all_leads'] = leads
        
        if "Follow Leaderboard" in wb.sheetnames:
            follow_sheet = wb["Follow Leaderboard"]
            follows = []
            for row in range(2, follow_sheet.max_row + 1):
                name = follow_sheet.cell(row=row, column=1).value
                points_cell = follow_sheet.cell(row=row, column=2).value
                if name:
                    # Try to parse points from the cell
                    points = 0
                    try:
                        if isinstance(points_cell, (int, float)):
                            points = int(points_cell)
                        elif isinstance(points_cell, str) and points_cell.isdigit():
                            points = int(points_cell)
                    except (ValueError, TypeError):
                        # If parsing fails, try to get points from our calculated data
                        if name in all_contestants['follows']:
                            points = all_contestants['follows'][name]['points']
                    
                    # Check if this contestant is a winner
                    is_winner = False
                    if name in all_contestants['follows']:
                        is_winner = all_contestants['follows'][name]['is_winner']
                    
                    follows.append({
                        'name': name,
                        'points': points,
                        'is_winner': is_winner
                    })
                    
                    # Update our tracking dictionary
                    if name in all_contestants['follows']:
                        all_contestants['follows'][name]['points'] = max(all_contestants['follows'][name]['points'], points)
            data['all_follows'] = follows
        
        # Convert tracking dictionaries to lists and use as final data if not already set
        if 'leads' not in data or not data['leads']:
            data['leads'] = list(all_contestants['leads'].values())
            # Sort by points in descending order
            data['leads'].sort(key=lambda x: x['points'], reverse=True)
            
            # Assign medals to top 3
            for i, lead in enumerate(data['leads'][:3]):
                if i == 0:
                    lead['medal'] = '🥇'
                elif i == 1:
                    lead['medal'] = '🥈'
                elif i == 2:
                    lead['medal'] = '🥉'
        
        if 'follows' not in data or not data['follows']:
            data['follows'] = list(all_contestants['follows'].values())
            # Sort by points in descending order
            data['follows'].sort(key=lambda x: x['points'], reverse=True)
            
            # Assign medals to top 3
            for i, follow in enumerate(data['follows'][:3]):
                if i == 0:
                    follow['medal'] = '🥇'
                elif i == 1:
                    follow['medal'] = '🥈'
                elif i == 2:
                    follow['medal'] = '🥉'
        
        # Clean up
        os.remove(temp_path)
        
        # Log data for debugging
        print("Processed data:", data)
        if 'leads' in data:
            for lead in data['leads']:
                print(f"Lead: {lead['name']}, Points: {lead['points']}, Type: {type(lead['points'])}")
        
        return jsonify(data)
    
    except Exception as e:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        print(f"Error processing Excel file: {str(e)}")
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

@app.route('/api/get_spotify_token', methods=['GET'])
def get_spotify_token():
    try:
        # Get client credentials from environment variables
        client_id = os.getenv('SPOTIFY_CLIENT_ID')
        client_secret = os.getenv('SPOTIFY_CLIENT_SECRET')
        
        if not client_id or not client_secret:
            return jsonify({'error': 'Spotify credentials not configured'}), 500
        
        # Request access token from Spotify
        auth_response = requests.post(
            'https://accounts.spotify.com/api/token',
            auth=(client_id, client_secret),
            data={'grant_type': 'client_credentials'}
        )
        
        if auth_response.status_code != 200:
            return jsonify({'error': 'Failed to get Spotify access token'}), 500
        
        return jsonify(auth_response.json())
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spotify/callback', methods=['GET'])
def spotify_callback():
    """Spotify OAuth callback endpoint."""
    code = request.args.get('code')
    state = request.args.get('state')
    error = request.args.get('error')

    if error:
        return f"Authentication failed: {error}"

    if not code or not state:
        return "Authentication failed: Missing code or state."

    try:
        # Decode the state parameter
        decoded_state = urllib.parse.unquote(state)
        state_data = json.loads(base64.b64decode(decoded_state).decode('utf-8'))
        session_id = state_data.get('session_id')
        return_to = state_data.get('return_to')
        auth_key = state_data.get('auth_key')

        if not session_id:
            return "Authentication failed: Invalid state."

        # Exchange the authorization code for an access token
        client_id = os.getenv('SPOTIFY_CLIENT_ID')
        client_secret = os.getenv('SPOTIFY_CLIENT_SECRET')
        if not client_id or not client_secret:
            return "Spotify credentials not configured."

        token_resp = requests.post(
            'https://accounts.spotify.com/api/token',
            auth=(client_id, client_secret),
            data={
                'grant_type': 'authorization_code',
                'code': code,
                'redirect_uri': _get_redirect_uri()
            }
        )
        if token_resp.status_code != 200:
            return f"Failed to exchange code for token: {token_resp.status_code} - {token_resp.text}"

        tokens = token_resp.json()
        access_token = tokens.get('access_token')
        refresh_token = tokens.get('refresh_token')
        expires_in = int(tokens.get('expires_in', 3600))

        if not access_token:
            return "Authentication failed: No access token received."

        # Store tokens using the resolved key
        _store_user_tokens(_resolve_key(session_id, auth_key), access_token, refresh_token, expires_in)

        # Redirect to the return_to URL if provided, otherwise to the game page
        if return_to:
            return redirect(return_to)
        else:
            return redirect(f"{request.host_url}?session_id={session_id}")
    except Exception as e:
        return f"Authentication failed: {str(e)}"

@app.route('/api/spotify/authorize', methods=['GET'])
def spotify_authorize():
    """Initiates Spotify OAuth flow."""
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400

    # Check if user is already authenticated for this session
    tokens = _get_user_tokens(_resolve_key(session_id, None)) # Pass None for auth_key
    if tokens and tokens.get('expires_at', 0) > _now():
        return jsonify({'message': 'User already authenticated for this session.'})

    redirect_uri = _get_redirect_uri()
    client_id = os.getenv('SPOTIFY_CLIENT_ID')
    if not client_id:
        return jsonify({'error': 'Spotify client ID not configured'}), 500

    scopes = [
        'user-read-private',
        'user-read-email',
        'user-read-playback-state',
        'user-modify-playback-state',
        'streaming'
    ]
    scope_str = ' '.join(scopes)

    # Encode state for redirect
    state_data = {
        'session_id': session_id,
        'return_to': request.args.get('return_to'), # Optional: redirect back to a specific page
        'auth_key': request.args.get('auth_key') # Optional: a unique key for this auth attempt
    }
    encoded_state = base64.b64encode(json.dumps(state_data).encode('utf-8')).decode('utf-8')

    params = {
        'response_type': 'code',
        'client_id': client_id,
        'redirect_uri': redirect_uri,
        'scope': scope_str,
        'state': encoded_state,
    }
    auth_url = 'https://accounts.spotify.com/authorize?' + urllib.parse.urlencode(params)
    return redirect(auth_url)

@app.route('/api/spotify/user_token', methods=['GET'])
def spotify_user_token():
    """Expose the current user's access token for the Web Playback SDK."""
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400
    access_token = _ensure_user_access_token(_resolve_key(session_id, None)) # Pass None for auth_key
    if not access_token:
        return jsonify({'error': 'Not authorized'}), 401
    tokens = _get_user_tokens(_resolve_key(session_id, None)) or {} # Pass None for auth_key
    return jsonify({
        'access_token': access_token,
        'expires_at': tokens.get('expires_at'),
    })

@app.route('/api/spotify/current_track', methods=['GET'])
def spotify_current_track():
    """Gets the currently playing track from the user's Spotify account."""
    session_id = request.args.get('session_id')
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400

    access_token = _ensure_user_access_token(_resolve_key(session_id, None)) # Pass None for auth_key
    if not access_token:
        return jsonify({'error': 'User not authenticated or token expired.'}), 401

    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        resp = requests.get('https://api.spotify.com/v1/me/player/currently-playing', headers=headers)
        if resp.status_code == 200:
            data = resp.json()
            item = data.get('item')
            if item:
                return jsonify({
                    'is_playing': data.get('is_playing'),
                    'track_name': item.get('name'),
                    'artist_name': ', '.join([a.get('name') for a in item.get('artists', []) if a]),
                    'spotify_url': item.get('external_urls', {}).get('spotify'),
                    'duration_ms': item.get('duration_ms'),
                    'progress_ms': data.get('progress_ms')
                })
            else:
                return jsonify({'is_playing': False, 'track_name': None, 'artist_name': None})
        elif resp.status_code == 401:
            # Token might be expired, try to refresh
            access_token = _refresh_user_token(_resolve_key(session_id, None)) # Pass None for auth_key
            if not access_token:
                return jsonify({'error': 'Failed to refresh Spotify token.'}), 500
            headers = {'Authorization': f'Bearer {access_token}'}
            resp = requests.get('https://api.spotify.com/v1/me/player/currently-playing', headers=headers)
            if resp.status_code == 200:
                data = resp.json()
                item = data.get('item')
                if item:
                    return jsonify({
                        'is_playing': data.get('is_playing'),
                        'track_name': item.get('name'),
                        'artist_name': ', '.join([a.get('name') for a in item.get('artists', []) if a]),
                        'spotify_url': item.get('external_urls', {}).get('spotify'),
                        'duration_ms': item.get('duration_ms'),
                        'progress_ms': data.get('progress_ms')
                    })
                else:
                    return jsonify({'is_playing': False, 'track_name': None, 'artist_name': None})
            else:
                return jsonify({'error': 'Failed to fetch currently playing track after refresh.', 'status': resp.status_code, 'details': resp.json()}), 502
        else:
            return jsonify({'error': 'Failed to fetch currently playing track', 'status': resp.status_code, 'details': resp.json()}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spotify/play_track', methods=['POST'])
def spotify_play_track():
    """Plays a track on the user's Spotify account."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON data provided'}), 400
    
    session_id = data.get('session_id')
    track_uri = data.get('track_uri')
    device_id = data.get('device_id')
    
    if not session_id or not track_uri:
        return jsonify({'error': 'Missing session_id or track_uri'}), 400

    access_token = _ensure_user_access_token(_resolve_key(session_id, None)) # Pass None for auth_key
    if not access_token:
        return jsonify({'error': 'User not authenticated or token expired.'}), 401

    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        url = 'https://api.spotify.com/v1/me/player/play'
        if device_id:
            url += f"?device_id={device_id}"
        resp = requests.put(url, headers=headers, json={'uris': [track_uri]})
        if resp.status_code == 204:
            return jsonify({'message': 'Track started successfully.'})
        elif resp.status_code == 401:
            access_token = _refresh_user_token(_resolve_key(session_id, None)) # Pass None for auth_key
            if not access_token:
                return jsonify({'error': 'Failed to refresh Spotify token.'}), 500
            headers = {'Authorization': f'Bearer {access_token}'}
            resp = requests.put(url, headers=headers, json={'uris': [track_uri]})
            if resp.status_code == 204:
                return jsonify({'message': 'Track started successfully after refresh.'})
            else:
                return jsonify({'error': 'Failed to start track after refresh.', 'status': resp.status_code, 'details': resp.json()}), 502
        else:
            return jsonify({'error': 'Failed to start track', 'status': resp.status_code, 'details': resp.json()}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spotify/pause_track', methods=['POST'])
def spotify_pause_track():
    """Pauses the user's Spotify account."""
    session_id = request.args.get('session_id')
    device_id = request.args.get('device_id')
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400

    access_token = _ensure_user_access_token(_resolve_key(session_id, None)) # Pass None for auth_key
    if not access_token:
        return jsonify({'error': 'User not authenticated or token expired.'}), 401

    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        url = 'https://api.spotify.com/v1/me/player/pause'
        if device_id:
            url += f"?device_id={device_id}"
        resp = requests.put(url, headers=headers)
        if resp.status_code == 204:
            return jsonify({'message': 'Track paused successfully.'})
        elif resp.status_code == 401:
            access_token = _refresh_user_token(_resolve_key(session_id, None)) # Pass None for auth_key
            if not access_token:
                return jsonify({'error': 'Failed to refresh Spotify token.'}), 500
            headers = {'Authorization': f'Bearer {access_token}'}
            resp = requests.put(url, headers=headers)
            if resp.status_code == 204:
                return jsonify({'message': 'Track paused successfully after refresh.'})
            else:
                return jsonify({'error': 'Failed to pause track after refresh.', 'status': resp.status_code, 'details': resp.json()}), 502
        else:
            return jsonify({'error': 'Failed to pause track', 'status': resp.status_code, 'details': resp.json()}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spotify/next_track', methods=['POST'])
def spotify_next_track():
    """Skips to the next track on the user's Spotify account."""
    session_id = request.args.get('session_id')
    device_id = request.args.get('device_id')
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400

    access_token = _ensure_user_access_token(_resolve_key(session_id, None)) # Pass None for auth_key
    if not access_token:
        return jsonify({'error': 'User not authenticated or token expired.'}), 401

    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        url = 'https://api.spotify.com/v1/me/player/next'
        if device_id:
            url += f"?device_id={device_id}"
        resp = requests.post(url, headers=headers)
        if resp.status_code == 204:
            return jsonify({'message': 'Track skipped successfully.'})
        elif resp.status_code == 401:
            access_token = _refresh_user_token(_resolve_key(session_id, None)) # Pass None for auth_key
            if not access_token:
                return jsonify({'error': 'Failed to refresh Spotify token.'}), 500
            headers = {'Authorization': f'Bearer {access_token}'}
            resp = requests.post(url, headers=headers)
            if resp.status_code == 204:
                return jsonify({'message': 'Track skipped successfully after refresh.'})
            else:
                return jsonify({'error': 'Failed to skip track after refresh.', 'status': resp.status_code, 'details': resp.json()}), 502
        else:
            return jsonify({'error': 'Failed to skip track', 'status': resp.status_code, 'details': resp.json()}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spotify/previous_track', methods=['POST'])
def spotify_previous_track():
    """Goes back to the previous track on the user's Spotify account."""
    session_id = request.args.get('session_id')
    device_id = request.args.get('device_id')
    if not session_id:
        return jsonify({'error': 'Missing session_id'}), 400

    access_token = _ensure_user_access_token(_resolve_key(session_id, None)) # Pass None for auth_key
    if not access_token:
        return jsonify({'error': 'User not authenticated or token expired.'}), 401

    try:
        headers = {'Authorization': f'Bearer {access_token}'}
        url = 'https://api.spotify.com/v1/me/player/previous'
        if device_id:
            url += f"?device_id={device_id}"
        resp = requests.post(url, headers=headers)
        if resp.status_code == 204:
            return jsonify({'message': 'Track went back successfully.'})
        elif resp.status_code == 401:
            access_token = _refresh_user_token(_resolve_key(session_id, None)) # Pass None for auth_key
            if not access_token:
                return jsonify({'error': 'Failed to refresh Spotify token.'}), 500
            headers = {'Authorization': f'Bearer {access_token}'}
            resp = requests.post(url, headers=headers)
            if resp.status_code == 204:
                return jsonify({'message': 'Track went back successfully after refresh.'})
            else:
                return jsonify({'error': 'Failed to go back track after refresh.', 'status': resp.status_code, 'details': resp.json()}), 502
        else:
            return jsonify({'error': 'Failed to go back track', 'status': resp.status_code, 'details': resp.json()}), 502
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/spotify/playlist_tracks', methods=['GET'])
def spotify_playlist_tracks():
    """Server-side proxy to fetch all tracks for a public Spotify playlist.
    Returns a simplified JSON structure: { tracks: [ { id, name, artists } ] }
    """
    try:
        playlist_id = request.args.get('playlist_id')
        if not playlist_id:
            return jsonify({'error': 'Missing playlist_id'}), 400

        client_id = os.getenv('SPOTIFY_CLIENT_ID')
        client_secret = os.getenv('SPOTIFY_CLIENT_SECRET')
        if not client_id or not client_secret:
            return jsonify({'error': 'Spotify credentials not configured'}), 500

        # Get access token
        token_resp = requests.post(
            'https://accounts.spotify.com/api/token',
            auth=(client_id, client_secret),
            data={'grant_type': 'client_credentials'}
        )
        if token_resp.status_code != 200:
            return jsonify({'error': 'Failed to get Spotify access token'}), 500
        access_token = token_resp.json().get('access_token')
        headers = {'Authorization': f'Bearer {access_token}'}

        # Fetch tracks (paginate)
        url = f'https://api.spotify.com/v1/playlists/{playlist_id}/tracks?limit=100'
        tracks = []
        visited = set()
        while url and url not in visited:
            visited.add(url)
            resp = requests.get(url, headers=headers)
            if resp.status_code == 401:
                # Try to refresh token once
                token_resp = requests.post(
                    'https://accounts.spotify.com/api/token',
                    auth=(client_id, client_secret),
                    data={'grant_type': 'client_credentials'}
                )
                if token_resp.status_code != 200:
                    return jsonify({'error': 'Failed to refresh Spotify token'}), 500
                access_token = token_resp.json().get('access_token')
                headers = {'Authorization': f'Bearer {access_token}'}
                resp = requests.get(url, headers=headers)
            if resp.status_code != 200:
                try:
                    return jsonify({'error': 'Failed to fetch playlist tracks', 'status': resp.status_code, 'details': resp.json()}), 502
                except Exception:
                    return jsonify({'error': 'Failed to fetch playlist tracks', 'status': resp.status_code}), 502
            data = resp.json()
            items = data.get('items', [])
            for item in items:
                t = (item or {}).get('track') or {}
                tid = t.get('id')
                if not tid:
                    continue
                tracks.append({
                    'id': tid,
                    'name': t.get('name', ''),
                    'artists': ', '.join([a.get('name', '') for a in t.get('artists', []) if a])
                })
            url = data.get('next')

        # Deduplicate by id
        unique = []
        seen = set()
        for t in tracks:
            if t['id'] in seen:
                continue
            seen.add(t['id'])
            unique.append(t)

        return jsonify({'tracks': unique})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host=config.HOST, port=config.PORT, debug=config.DEBUG) 