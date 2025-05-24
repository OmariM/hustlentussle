import unittest
import io
import sys
import os
from openpyxl import load_workbook
from web.app import app, games
from game_logic import Game
from web.config import get_config

class TestGameLogic(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """Set up test environment once for all tests."""
        # Get test configuration
        config = get_config()
        config.PORT = 5001  # Use a different port for testing
        config.DEBUG = True
        config.TESTING = True
        
        # Configure the app
        app.config.from_object(config)
    
    def setUp(self):
        """Set up test environment before each test."""
        self.app_context = app.app_context()
        self.app_context.push()
        self.client = app.test_client()
        
        # Clear any existing games
        games.clear()
        
        # Create test data
        self.lead_names = ["Lead1", "Lead2", "Lead3", "Lead4"]
        self.follow_names = ["Follow1", "Follow2", "Follow3", "Follow4"]
        self.judge_names = ["Judge1", "Judge2"]
        
        # Create a test game
        self.game = Game(self.lead_names, self.follow_names, self.judge_names)
    
    def tearDown(self):
        """Clean up after each test."""
        try:
            self.app_context.pop()
            games.clear()
        except Exception as e:
            print(f"Error in tearDown: {str(e)}")
    
    def simulate_round(self, pair, role, votes):
        """Helper method to simulate a round."""
        return self.game.judge_round(pair[0], pair[1], role, votes)
    
    def test_tie_on_leads(self):
        """Test that a tie is correctly handled for leads."""
        lead_pair = (self.game.pair_1[0], self.game.pair_2[0])
        result = self.simulate_round(
            lead_pair, "lead", 
            [("Judge1", 3), ("Judge2", 3), ("Lead3", 1), ("Lead4", 2)]
        )
        self.assertTrue(
            result["winner"].startswith("Tie between"),
            f"Expected tie, got: {result['winner']}"
        )
    
    def test_no_contest_on_follows(self):
        """Test that a no contest is correctly handled for follows."""
        follow_pair = (self.game.pair_1[1], self.game.pair_2[1])
        result = self.simulate_round(
            follow_pair, "follow",
            [("Judge1", 4), ("Judge2", 4), ("Follow3", 1), ("Follow4", 2)]
        )
        self.assertEqual(
            result["winner"], "No Contest",
            f"Expected 'No Contest', got: {result['winner']}"
        )
    
    def test_split_vote_on_leads(self):
        """Test that split votes are correctly handled for leads."""
        lead_pair = (self.game.pair_1[0], self.game.pair_2[0])
        result = self.simulate_round(
            lead_pair, "lead",
            [("Judge1", 1), ("Judge2", 2), ("Lead3", 1), ("Lead4", 2)]
        )
        self.assertEqual(
            result["winner"], lead_pair[0].name,
            f"Expected {lead_pair[0].name}, got: {result['winner']}"
        )
    
    def test_sweep_on_follows(self):
        """Test that a sweep vote is correctly handled for follows."""
        follow_pair = (self.game.pair_1[1], self.game.pair_2[1])
        result = self.simulate_round(
            follow_pair, "follow",
            [("Judge1", 1), ("Judge2", 1), ("Follow3", 1), ("Follow4", 1)]
        )
        self.assertEqual(
            result["winner"], follow_pair[0].name,
            f"Expected {follow_pair[0].name}, got: {result['winner']}"
        )
    
    def test_double_tie_on_leads(self):
        """Test that double ties are correctly handled for leads."""
        lead_pair = (self.game.pair_1[0], self.game.pair_2[0])
        # First tie
        result = self.simulate_round(
            lead_pair, "lead",
            [("Judge1", 3), ("Judge2", 3), ("Lead3", 1), ("Lead4", 2)]
        )
        self.assertTrue(
            result["winner"].startswith("Tie between"),
            f"Expected tie, got: {result['winner']}"
        )
        
        # Move to next round and verify tied leads are re-paired
        self.game.next_round()
        tied_names = {lead_pair[0].name, lead_pair[1].name}
        next_leads = {self.game.pair_1[0].name, self.game.pair_2[0].name}
        self.assertEqual(
            tied_names, next_leads,
            f"Expected re-paired leads {tied_names}, got {next_leads}"
        )
    
    def test_no_contest_queue_placement(self):
        """Test that no contest contestants are placed at the end of the queue."""
        lead_pair = (self.game.pair_1[0], self.game.pair_2[0])
        self.simulate_round(
            lead_pair, "lead",
            [("Judge1", 4), ("Judge2", 4), ("Lead3", 1), ("Lead4", 2)]
        )
        after_leads = [c.name for c in self.game.leads]
        expected_end = [lead_pair[0].name, lead_pair[1].name]
        self.assertEqual(
            after_leads[-2:], expected_end,
            f"Expected end {expected_end}, got {after_leads[-2:]}"
        )
    
    def test_early_exit_finalization(self):
        """Test that early exit correctly finalizes and sorts results."""
        lead_pair = (self.game.pair_1[0], self.game.pair_2[0])
        follow_pair = (self.game.pair_1[1], self.game.pair_2[1])
        
        # Simulate one round of clear wins
        self.simulate_round(
            lead_pair, "lead",
            [("Judge1", 1), ("Judge2", 1), ("Lead3", 1), ("Lead4", 1)]
        )
        self.simulate_round(
            follow_pair, "follow",
            [("Judge1", 1), ("Judge2", 1), ("Follow3", 1), ("Follow4", 1)]
        )
        
        # Early exit
        leads_sorted, follows_sorted = self.game.finalize_results()
        self.assertGreater(
            leads_sorted[0].points, leads_sorted[1].points,
            "Top lead not correctly sorted"
        )
        self.assertGreater(
            follows_sorted[0].points, follows_sorted[1].points,
            "Top follow not correctly sorted"
        )
    
    def test_lead_win_condition(self):
        """Test that lead win condition is correctly handled."""
        game = Game(
            ["Lead1", "Lead2", "Lead3", "Lead4", "Lead5"],
            ["Follow1", "Follow2", "Follow3", "Follow4", "Follow5"],
            ["Judge1", "Judge2"]
        )
        
        initial_lead_1 = game.pair_1[0]
        initial_lead_2 = game.pair_2[0]
        
        # Give winning points to Lead1
        initial_lead_1.points = game.total_num_leads - 2
        
        # Simulate round where Lead1 wins
        lead_pair = (initial_lead_1, initial_lead_2)
        result = game.judge_round(
            lead_pair[0], lead_pair[1], "lead",
            [("Judge1", 1), ("Judge2", 1)]
        )
        
        # Check if Lead1 is marked as a winner
        self.assertTrue(
            game.has_winning_lead and game.winning_lead.name == initial_lead_1.name,
            f"Lead winner not correctly identified. Has winning lead: {game.has_winning_lead}, "
            f"Winning lead: {game.winning_lead.name if game.winning_lead else None}"
        )
        
        # Move to next round
        game.next_round()
        
        # Check if the winning lead is in the queue
        leads_in_queue = [lead.name for lead in game.leads]
        self.assertIn(
            initial_lead_1.name, leads_in_queue,
            f"Winning lead not in queue. Queue: {leads_in_queue}"
        )
        
        # Check that two different leads are now competing
        competing_leads = [game.pair_1[0].name, game.pair_2[0].name]
        self.assertNotIn(
            initial_lead_1.name, competing_leads,
            f"Winning lead still competing. Current competitors: {competing_leads}"
        )

    def test_tie_on_follows(self):
        """Test that when there's a tie between follows, the same follows compete again with the winning lead and next lead from queue."""
        # Get initial pairs
        initial_pair1 = (self.game.pair_1[0].name, self.game.pair_1[1].name)
        initial_pair2 = (self.game.pair_2[0].name, self.game.pair_2[1].name)
        
        # First simulate a lead round to establish a winning lead
        lead_pair = (self.game.pair_1[0], self.game.pair_2[0])
        lead_result = self.simulate_round(
            lead_pair, "lead",
            [("Judge1", 1), ("Judge2", 1), ("Lead3", 1), ("Lead4", 1)]
        )
        
        # Verify lead winner was recorded
        self.assertEqual(
            lead_result["winner"], lead_pair[0].name,
            f"Expected {lead_pair[0].name} to win lead round, got: {lead_result['winner']}"
        )
        
        # Store the next lead from queue before the follow tie
        next_lead = self.game.leads[0]
        
        # Now simulate a tie between follows
        follow_pair = (self.game.pair_1[1], self.game.pair_2[1])
        follow_result = self.simulate_round(
            follow_pair, "follow",
            [("Judge1", 3), ("Judge2", 3), ("Follow3", 1), ("Follow4", 2)]
        )
        
        # Verify tie was recorded
        self.assertTrue(
            follow_result["winner"].startswith("Tie between"),
            f"Expected tie, got: {follow_result['winner']}"
        )
        
        # Move to next round
        self.game.next_round()
        
        # Get new pairs
        new_pair1 = (self.game.pair_1[0].name, self.game.pair_1[1].name)
        new_pair2 = (self.game.pair_2[0].name, self.game.pair_2[1].name)
        
        # Verify that the follows are the same as before
        self.assertEqual(
            {initial_pair1[1], initial_pair2[1]},
            {new_pair1[1], new_pair2[1]},
            "Follows should remain the same after follow tie"
        )
        
        # Verify that one lead is the winning lead from previous round
        self.assertIn(
            self.game.last_lead_winner,
            [new_pair1[0], new_pair2[0]],
            "Winning lead should stay in competition"
        )
        
        # Verify that the other lead is the next lead from queue
        self.assertIn(
            next_lead.name,
            [new_pair1[0], new_pair2[0]],
            "Next lead from queue should be competing"
        )
        
        # Verify that no couple is the same as before
        self.assertFalse(
            (new_pair1 == initial_pair1 or new_pair1 == initial_pair2 or
             new_pair2 == initial_pair1 or new_pair2 == initial_pair2),
            "No couple should remain the same after follow tie"
        )

class TestExportBattleData(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        """Set up test environment once for all tests."""
        # Get test configuration
        config = get_config()
        config.PORT = 5001  # Use a different port for testing
        config.DEBUG = True
        config.TESTING = True
        
        # Configure the app
        app.config.from_object(config)
    
    def setUp(self):
        """Set up test environment before each test."""
        self.app_context = app.app_context()
        self.app_context.push()
        self.client = app.test_client()
        
        # Clear any existing games
        games.clear()
        
        # Create a test game with known initial order
        self.lead_names = ["Lead1", "Lead2", "Lead3", "Lead4"]
        self.follow_names = ["Follow1", "Follow2", "Follow3", "Follow4"]
        self.judge_names = ["Judge1", "Judge2"]
        
        try:
            # Start a new game
            response = self.client.post('/api/start_game', json={
                'leads': ','.join(self.lead_names),
                'follows': ','.join(self.follow_names),
                'judges': ','.join(self.judge_names)
            })
            
            if response.status_code != 200:
                print(f"Error starting game: {response.get_data(as_text=True)}")
                self.fail("Failed to start game")
            
            self.session_id = response.json['session_id']
            
            # Store the initial order from the response
            self.initial_leads = response.json['initial_leads']
            self.initial_follows = response.json['initial_follows']
            
            # Add a test round
            self.add_test_round()
            
        except Exception as e:
            print(f"Error in setUp: {str(e)}")
            self.fail(f"Setup failed: {str(e)}")
    
    def tearDown(self):
        """Clean up after each test."""
        try:
            self.app_context.pop()
            games.clear()
        except Exception as e:
            print(f"Error in tearDown: {str(e)}")
    
    def add_test_round(self):
        """Add a test round with votes."""
        try:
            # Get current pairs
            response = self.client.get(f'/api/get_scores?session_id={self.session_id}')
            if response.status_code != 200:
                print(f"Error getting scores: {response.get_data(as_text=True)}")
                self.fail("Failed to get scores")
            
            game = games[self.session_id]
            
            # Add lead votes
            lead_response = self.client.post('/api/judge_leads', json={
                'session_id': self.session_id,
                'votes': [
                    {'judge': 'Judge1', 'vote': 1},
                    {'judge': 'Judge2', 'vote': 1}
                ],
                'song_info': {
                    'title': 'Test Song',
                    'artist': 'Test Artist',
                    'spotify_url': 'https://spotify.com/test'
                }
            })
            
            if lead_response.status_code != 200:
                print(f"Error adding lead votes: {lead_response.get_data(as_text=True)}")
                self.fail("Failed to add lead votes")
            
            # Add follow votes
            follow_response = self.client.post('/api/judge_follows', json={
                'session_id': self.session_id,
                'votes': [
                    {'judge': 'Judge1', 'vote': 2},
                    {'judge': 'Judge2', 'vote': 2}
                ]
            })
            
            if follow_response.status_code != 200:
                print(f"Error adding follow votes: {follow_response.get_data(as_text=True)}")
                self.fail("Failed to add follow votes")
                
        except Exception as e:
            print(f"Error in add_test_round: {str(e)}")
            self.fail(f"Failed to add test round: {str(e)}")
    
    def test_initial_order_export(self):
        """Test that the initial order is correctly exported in the Excel file."""
        try:
            # Export battle data
            response = self.client.get(f'/api/export_battle_data?session_id={self.session_id}')
            
            # Verify response
            self.assertEqual(response.status_code, 200, 
                           f"Export failed with status {response.status_code}")
            self.assertEqual(response.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           "Incorrect mimetype for Excel file")
            
            # Load the Excel file from the response
            excel_data = io.BytesIO(response.data)
            try:
                wb = load_workbook(excel_data)
            except Exception as e:
                print(f"Error loading Excel file: {str(e)}")
                self.fail("Failed to load Excel file")
            
            # Get the Battle Summary sheet
            self.assertIn("Battle Summary", wb.sheetnames, "Battle Summary sheet not found")
            summary_sheet = wb["Battle Summary"]
            
            # Find the initial order section
            current_row = 7  # Start after "Initial Order" header
            found_initial_order = False
            
            # First, find the "Initial Order" section
            while current_row < summary_sheet.max_row:
                cell_value = summary_sheet[f'A{current_row}'].value
                if cell_value == "Initial Order":
                    found_initial_order = True
                    current_row += 2  # Skip the header and move to content
                    break
                current_row += 1
            
            self.assertTrue(found_initial_order, "Initial Order section not found in Excel file")
            
            # Extract leads from Excel
            excel_leads = []
            while current_row < summary_sheet.max_row:
                lead_entry = summary_sheet[f'A{current_row}'].value
                if not lead_entry or lead_entry == "Follows:":
                    break
                if isinstance(lead_entry, str) and '. ' in lead_entry:
                    lead_name = lead_entry.split('. ', 1)[1]
                    excel_leads.append(lead_name)
                current_row += 1
            
            # Extract follows from Excel
            excel_follows = []
            # Skip the "Follows:" header
            current_row += 1
            while current_row < summary_sheet.max_row:
                follow_entry = summary_sheet[f'A{current_row}'].value
                if not follow_entry or follow_entry == "Final Results":
                    break
                if isinstance(follow_entry, str) and '. ' in follow_entry:
                    follow_name = follow_entry.split('. ', 1)[1]
                    excel_follows.append(follow_name)
                current_row += 1
            
            # Verify that the exported order matches the initial order
            self.assertEqual(excel_leads, self.initial_leads, 
                           f"Exported leads order does not match initial order.\nExpected: {self.initial_leads}\nGot: {excel_leads}")
            self.assertEqual(excel_follows, self.initial_follows, 
                           f"Exported follows order does not match initial order.\nExpected: {self.initial_follows}\nGot: {excel_follows}")
            
        except Exception as e:
            print(f"Error in test_initial_order_export: {str(e)}")
            self.fail(f"Test failed: {str(e)}")

    def test_song_info_export(self):
        """Test that song information is correctly extracted and exported."""
        try:
            # Add a test round with song information
            lead_response = self.client.post('/api/judge_leads', json={
                'session_id': self.session_id,
                'votes': [
                    {'judge': 'Judge1', 'vote': 1},
                    {'judge': 'Judge2', 'vote': 1}
                ],
                'song_info': {
                    'title': 'Test Song Title',
                    'artist': 'Test Artist Name',
                    'spotify_url': 'https://open.spotify.com/track/1234567890'
                }
            })
            
            if lead_response.status_code != 200:
                print(f"Error adding lead votes: {lead_response.get_data(as_text=True)}")
                self.fail("Failed to add lead votes")
            
            # Add follow votes to complete the round
            follow_response = self.client.post('/api/judge_follows', json={
                'session_id': self.session_id,
                'votes': [
                    {'judge': 'Judge1', 'vote': 2},
                    {'judge': 'Judge2', 'vote': 2}
                ]
            })
            
            if follow_response.status_code != 200:
                print(f"Error adding follow votes: {follow_response.get_data(as_text=True)}")
                self.fail("Failed to add follow votes")
            
            # Export battle data
            response = self.client.get(f'/api/export_battle_data?session_id={self.session_id}')
            
            # Verify response
            self.assertEqual(response.status_code, 200, 
                           f"Export failed with status {response.status_code}")
            
            # Load the Excel file
            excel_data = io.BytesIO(response.data)
            wb = load_workbook(excel_data)
            
            # Get the Round History sheet
            self.assertIn("Round History", wb.sheetnames, "Round History sheet not found")
            round_sheet = wb["Round History"]
            
            # Find the row with our test round
            found_round = False
            for row in range(2, round_sheet.max_row + 1):
                if round_sheet.cell(row=row, column=1).value == 1:  # Round 1
                    found_round = True
                    # Check song information
                    song_title = round_sheet.cell(row=row, column=6).value
                    artist = round_sheet.cell(row=row, column=7).value
                    spotify_url = round_sheet.cell(row=row, column=8).value
                    
                    self.assertEqual(song_title, 'Test Song Title', 
                                   f"Song title mismatch. Expected: Test Song Title, Got: {song_title}")
                    self.assertEqual(artist, 'Test Artist Name', 
                                   f"Artist name mismatch. Expected: Test Artist Name, Got: {artist}")
                    self.assertEqual(spotify_url, 'https://open.spotify.com/track/1234567890', 
                                   f"Spotify URL mismatch. Expected: https://open.spotify.com/track/1234567890, Got: {spotify_url}")
                    break
            
            self.assertTrue(found_round, "Test round not found in exported data")
            
        except Exception as e:
            print(f"Error in test_song_info_export: {str(e)}")
            self.fail(f"Test failed: {str(e)}")

if __name__ == '__main__':
    unittest.main() 